#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
苏超智能化指挥中心 - 数据血缘表 V6
消除BFF中台数据联表的4个架构隐患盲区
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

INPUT_FILE = '苏超智能化指挥中心_数据血缘表_V5_按数据源重组版.xlsx'
OUTPUT_FILE = '苏超智能化指挥中心_数据血缘表_V6_消除盲区版.xlsx'

# ==================== 盲区 1: MAE_PERF_15MIN 新增 VIP 代理指标 ====================
def fix_blind_spot_1_mae_vip(ws):
    """在MAE表最后新增vip_rrc_users行"""
    print("\n[盲区1] 处理 MAE_PERF_15MIN - 新增VIP代理指标...")
    
    # 找到当前最后一行的行号
    max_row = ws.max_row
    
    # 新增行的数据
    new_row = [
        max_row - 6 + 1,  # 序号 (从第7行开始是数据，所以减去6)
        'vip_rrc_users',
        'Integer',
        '高优RRC连接数(网管侧VIP代理指标，极重要：必须根据 5QI=6 或特定 SPID 过滤)',
        '基于 RRC.ConnMax 增加 5QI 过滤维度'
    ]
    
    # 追加到最后一行
    ws.append(new_row)
    print(f"  已新增行: vip_rrc_users (序号 {new_row[0]})")
    
    return ws

# ==================== 盲区 2: MANUAL_CELL_CONFIG zone_name 强化 ====================
def fix_blind_spot_2_zone_mapping(ws):
    """强化zone_name字段的3D空间映射说明"""
    print("\n[盲区2] 处理 MANUAL_CELL_CONFIG - 强化zone_name字段...")
    
    # 遍历所有行，找到字段名为 zone_name 的那一行
    for row_idx in range(7, ws.max_row + 1):  # 从第7行开始(表头在第6行)
        cell_value = ws.cell(row=row_idx, column=2).value  # 第2列是字段名
        if cell_value and cell_value.strip() == 'zone_name':
            # 更新第4列(字段说明)
            new_comment = '场内防线区(如南看台，极其重要！这是 BFF 中台进行 3D 空间地图映射的唯一桥梁)'
            ws.cell(row=row_idx, column=4).value = new_comment
            print(f"  已更新第 {row_idx} 行 zone_name 字段说明")
            break
    
    return ws

# ==================== 盲区 3: DSP_KQI_15MIN 枚举值强约束 ====================
def fix_blind_spot_3_dsp_constraint(ws):
    """在DSP表的业务指标行追加契约约束"""
    print("\n[盲区3] 处理 DSP_KQI_15MIN - 追加契约约束...")
    
    constraint_text = ' [契约强约束：JSON 返回的 Key 必须严格遵守此字段名，严禁随意更改拼写]'
    
    # 遍历第4-10行（业务指标行）
    for row_idx in range(4 + 6, 10 + 6 + 1):  # +6 是因为数据从第7行开始
        if row_idx > ws.max_row:
            break
        
        # 获取当前字段说明
        current_comment = ws.cell(row=row_idx, column=4).value
        if current_comment:
            # 追加约束文本
            ws.cell(row=row_idx, column=4).value = str(current_comment) + constraint_text
            field_name = ws.cell(row=row_idx, column=2).value
            print(f"  第 {row_idx - 6} 行 ({field_name}) 已追加约束")
    
    return ws

# ==================== 盲区 4: 时间戳格式全局对齐 ====================
def fix_blind_spot_4_timestamp_global(wb):
    """全局时间戳格式对齐"""
    print("\n[盲区4] 全局时间戳格式对齐...")
    
    timestamp_suffix = ' (必须为 13位毫秒级 Unix Timestamp)'
    iron_rule_text = '所有底层表的时间戳字段(timestamp)，必须统一采用 13位毫秒级 Unix Epoch 时间戳（如 1709542800000），严禁使用带时区的字符串！'
    
    # 处理 00_总目录表 - 新增列
    print("  处理 00_总目录表...")
    ws_toc = wb['00_总目录']
    
    # 找到最后一列的列号
    max_col = ws_toc.max_column
    new_col = max_col + 1
    
    # 新增列标题
    ws_toc.cell(row=1, column=new_col).value = '全局研发铁律'
    
    # 设置表头样式
    ws_toc.cell(row=1, column=new_col).fill = PatternFill(
        start_color='FF0000', end_color='FF0000', fill_type='solid'
    )
    ws_toc.cell(row=1, column=new_col).font = Font(bold=True, color='FFFFFF')
    
    # 在第一行数据填入铁律
    ws_toc.cell(row=2, column=new_col).value = iron_rule_text
    print(f"  已在 00_总目录 新增列 '全局研发铁律'")
    
    # 处理所有明细表
    skip_sheets = ['00_总目录', '旧新表映射关系']
    
    for sheet_name in wb.sheetnames:
        if sheet_name in skip_sheets:
            continue
        
        ws = wb[sheet_name]
        print(f"  处理 {sheet_name}...")
        
        # 找到字段名为 timestamp 的行
        for row_idx in range(7, ws.max_row + 1):
            field_name = ws.cell(row=row_idx, column=2).value
            if field_name and field_name.strip() == 'timestamp':
                # 获取当前字段说明
                current_comment = ws.cell(row=row_idx, column=4).value
                if current_comment:
                    # 追加时间戳格式要求
                    new_comment = str(current_comment) + timestamp_suffix
                    ws.cell(row=row_idx, column=4).value = new_comment
                    print(f"    已更新 timestamp 字段说明")
                break
    
    return wb

# ==================== 主程序 ====================
def main():
    print("=" * 80)
    print("苏超智能化指挥中心 - 数据血缘表 V6")
    print("消除BFF中台数据联表的4个架构隐患盲区")
    print("=" * 80)
    
    # 加载工作簿
    print(f"\n加载输入文件: {INPUT_FILE}")
    wb = load_workbook(INPUT_FILE)
    
    # 盲区1: MAE_PERF_15MIN
    if 'MAE_PERF_15MIN' in wb.sheetnames:
        ws_mae = wb['MAE_PERF_15MIN']
        fix_blind_spot_1_mae_vip(ws_mae)
    
    # 盲区2: MANUAL_CELL_CONFIG
    if 'MANUAL_CELL_CONFIG' in wb.sheetnames:
        ws_manual = wb['MANUAL_CELL_CONFIG']
        fix_blind_spot_2_zone_mapping(ws_manual)
    
    # 盲区3: DSP_KQI_15MIN
    if 'DSP_KQI_15MIN' in wb.sheetnames:
        ws_dsp = wb['DSP_KQI_15MIN']
        fix_blind_spot_3_dsp_constraint(ws_dsp)
    
    # 盲区4: 全局时间戳对齐
    wb = fix_blind_spot_4_timestamp_global(wb)
    
    # 保存
    print(f"\n保存输出文件: {OUTPUT_FILE}")
    wb.save(OUTPUT_FILE)
    
    print("\n" + "=" * 80)
    print("✅ 所有盲区已消除！")
    print("=" * 80)

if __name__ == '__main__':
    main()
