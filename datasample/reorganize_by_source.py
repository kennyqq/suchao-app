#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
苏超智能化指挥中心 - 数据血缘表 V5
按数据源系统 + 时间颗粒度重新切分底层表格
"""

import pandas as pd
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

INPUT_FILE = '数据血缘表-v4-最终版.xlsx'
OUTPUT_FILE = '苏超智能化指挥中心_数据血缘表_V5_按数据源重组版.xlsx'

# ==================== 定义新的表结构 ====================

# 新表1: MAE_15分钟_基站性能 (无线网络管平台)
MAE_15MIN_PERF = {
    'table_name': 'MAE_PERF_15MIN',
    'table_comment': '无线网络管平台-基站性能指标(15分钟粒度)',
    'source_system': '无线网络管平台 (MAE)',
    'granularity': '15分钟',
    'integration_method': '北向接口/CSV',
    'fields': [
        {'name': 'timestamp', 'type': 'Long', 'comment': '时间戳', 'nms_field': ''},
        {'name': 'cell_id', 'type': 'String', 'comment': '小区ID', 'nms_field': ''},
        {'name': 'rrc_conn_users', 'type': 'Integer', 'comment': 'RRC连接用户数', 'nms_field': 'RRC.ConnMax'},
        {'name': 'prb_util_ul', 'type': 'Float', 'comment': '上行PRB利用率', 'nms_field': 'RRU.PuschPrbAssn'},
        {'name': 'prb_util_dl', 'type': 'Float', 'comment': '下行PRB利用率', 'nms_field': 'RRU.PdschPrbAssn'},
        {'name': 'traffic_total_ul_mb', 'type': 'Float', 'comment': '上行流量(MB)', 'nms_field': 'MAC.CpOctUl'},
        {'name': 'traffic_total_dl_mb', 'type': 'Float', 'comment': '下行流量(MB)', 'nms_field': 'MAC.CpOctDl'},
    ]
}

# 新表2: DSP_15分钟_KQI业务质量 (智能板)
DSP_15MIN_KQI = {
    'table_name': 'DSP_KQI_15MIN',
    'table_comment': '无线智能板-KQI业务质量指标(15分钟粒度)',
    'source_system': '无线智能板 (DSP)',
    'granularity': '15分钟/实时',
    'integration_method': 'API接口',
    'fields': [
        {'name': 'timestamp', 'type': 'Long', 'comment': '时间戳', 'nms_field': ''},
        {'name': 'cell_id', 'type': 'String', 'comment': '小区ID', 'nms_field': ''},
        {'name': 'ue_5ga_ratio', 'type': 'Float', 'comment': '5G-A终端渗透率', 'nms_field': ''},
        {'name': 'dy_video_first_frame_delay_ms', 'type': 'Float', 'comment': '抖音首帧时延(ms)', 'nms_field': 'SVid.TotVidInitDlyTimeDl'},
        {'name': 'dy_video_freeze_rate', 'type': 'Float', 'comment': '抖音卡顿率', 'nms_field': 'SVid.TotVidPauseDur'},
        {'name': 'wx_msg_success_rate', 'type': 'Float', 'comment': '微信消息成功率', 'nms_field': ''},
        {'name': 'wx_pic_ul_rate_mbps', 'type': 'Float', 'comment': '微信图片上传速率(Mbps)', 'nms_field': ''},
        {'name': 'game_avg_delay_ms', 'type': 'Float', 'comment': '游戏平均时延(ms)', 'nms_field': 'SVid.TotDlVidTcpRtt'},
        {'name': 'pay_scan_delay_ms', 'type': 'Float', 'comment': '扫码支付时延(ms)', 'nms_field': ''},
        {'name': 'live_hd_ul_peak_rate_mbps', 'type': 'Float', 'comment': '直播上行峰值速率(Mbps)', 'nms_field': ''},
    ]
}

# 新表3: SEQ_小时级_用户数据 (共享层)
SEQ_HOURLY_USER = {
    'table_name': 'SEQ_USER_HOURLY',
    'table_comment': 'SEQ系统(共享层)-用户级数据(小时粒度)',
    'source_system': 'SEQ系统(共享层)',
    'granularity': '小时级',
    'integration_method': 'API接口/FTP',
    'fields': [
        {'name': 'timestamp', 'type': 'Long', 'comment': '时间戳(小时)', 'nms_field': ''},
        {'name': 'user_id', 'type': 'String', 'comment': '用户伪码', 'nms_field': ''},
        {'name': 'current_poi', 'type': 'String', 'comment': '当前驻留POI', 'nms_field': ''},
        {'name': 'home_city', 'type': 'String', 'comment': '归属城市', 'nms_field': ''},
        {'name': 'user_type', 'type': 'String', 'comment': '用户类型(resident/visitor)', 'nms_field': ''},
        {'name': 'stay_duration', 'type': 'Integer', 'comment': '驻留时长(分钟)', 'nms_field': ''},
    ]
}

# 新表4: SEQ_天级_用户画像 (共享层)
SEQ_DAILY_USER_PROFILE = {
    'table_name': 'SEQ_USER_PROFILE_DAILY',
    'table_comment': 'SEQ系统(共享层)-用户画像维表(天级T+1)',
    'source_system': 'SEQ系统(共享层)',
    'granularity': '天级(T+1)',
    'integration_method': 'API接口/FTP',
    'fields': [
        {'name': 'user_id', 'type': 'String', 'comment': '用户伪码(主键)', 'nms_field': ''},
        {'name': 'user_type', 'type': 'String', 'comment': '宏观属性(resident/visitor)', 'nms_field': ''},
        {'name': 'home_city', 'type': 'String', 'comment': '归属城市', 'nms_field': ''},
        {'name': 'user_tier', 'type': 'String', 'comment': '客户层级(vip/normal)', 'nms_field': ''},
        {'name': 'ue_tac_model', 'type': 'String', 'comment': '终端型号', 'nms_field': ''},
        {'name': 'ue_5ga_capable', 'type': 'Boolean', 'comment': '5G-A能力', 'nms_field': ''},
    ]
}

# 新表5: SEQ_小时级_终端统计 (共享层)
SEQ_HOURLY_DEVICE = {
    'table_name': 'SEQ_DEVICE_HOURLY',
    'table_comment': 'SEQ系统(共享层)-终端统计(小时粒度)',
    'source_system': 'SEQ系统(共享层)',
    'granularity': '小时级',
    'integration_method': 'API',
    'fields': [
        {'name': 'timestamp', 'type': 'Long', 'comment': '时间戳(小时)', 'nms_field': ''},
        {'name': 'zone_name', 'type': 'String', 'comment': '防线区', 'nms_field': ''},
        {'name': 'top_devices_json', 'type': 'JSON', 'comment': '热门终端TOP5', 'nms_field': ''},
    ]
}

# 新表6: AUTIN_实时_告警事件
AUTIN_REALTIME_ALARM = {
    'table_name': 'AUTIN_ALARM_REALTIME',
    'table_comment': '故障告警中心(AUTIN)-告警事件流(实时)',
    'source_system': '故障告警中心 (AUTIN)',
    'granularity': '实时/秒级',
    'integration_method': 'WebSocket推送',
    'fields': [
        {'name': 'event_id', 'type': 'String', 'comment': '事件流水号(主键)', 'nms_field': ''},
        {'name': 'timestamp', 'type': 'Long', 'comment': '发生绝对时间戳', 'nms_field': ''},
        {'name': 'alarm_level', 'type': 'Integer', 'comment': '告警级别(1高危2中3低)', 'nms_field': ''},
        {'name': 'alarm_title', 'type': 'String', 'comment': '告警标题', 'nms_field': ''},
        {'name': 'cell_id', 'type': 'String', 'comment': '关联小区ID', 'nms_field': ''},
        {'name': 'ai_root_cause_diagnosis', 'type': 'String', 'comment': 'AI根因诊断', 'nms_field': ''},
    ]
}

# 新表7: 人工_静态_网元配置
MANUAL_STATIC_CONFIG = {
    'table_name': 'MANUAL_CELL_CONFIG',
    'table_comment': '人工维护-网元公参配置表(静态)',
    'source_system': '人工维护/网优团队',
    'granularity': '周级',
    'integration_method': 'Excel/CSV导入',
    'fields': [
        {'name': 'cell_id', 'type': 'String', 'comment': '小区ID(主键)', 'nms_field': ''},
        {'name': 'cell_name', 'type': 'String', 'comment': '小区名称', 'nms_field': ''},
        {'name': 'poi_name', 'type': 'String', 'comment': '宏观POI归属地', 'nms_field': ''},
        {'name': 'zone_name', 'type': 'String', 'comment': '场内防线区', 'nms_field': ''},
        {'name': 'lng', 'type': 'Float', 'comment': '经度(高德)', 'nms_field': ''},
        {'name': 'lat', 'type': 'Float', 'comment': '纬度(高德)', 'nms_field': ''},
        {'name': 'cell_type', 'type': 'String', 'comment': '网元类型(macro/vehicle/micro)', 'nms_field': ''},
        {'name': 'is_3cc', 'type': 'Boolean', 'comment': '3CC载波支持', 'nms_field': ''},
        {'name': 'is_5ga', 'type': 'Boolean', 'comment': '是否5G-A小区', 'nms_field': ''},
        {'name': 'has_smart_board', 'type': 'Boolean', 'comment': '智能板状态', 'nms_field': ''},
        {'name': 'baseline_rrc_users', 'type': 'Integer', 'comment': '平日同时段基准人数', 'nms_field': ''},
    ]
}

# 新表8: 人工_静态_容量配置
MANUAL_STATIC_CAPACITY = {
    'table_name': 'MANUAL_CAPACITY_CONFIG',
    'table_comment': '人工维护-区域容量配置表(静态)',
    'source_system': '人工维护',
    'granularity': '小时级',
    'integration_method': '人工表单接口',
    'fields': [
        {'name': 'zone_name', 'type': 'String', 'comment': '防线区(主键)', 'nms_field': ''},
        {'name': 'cell_capacity_config', 'type': 'Integer', 'comment': '区域网络容量配置', 'nms_field': ''},
    ]
}

# 新表9: 人工_赛后单次_复盘数据
MANUAL_POST_MATCH_SUMMARY = {
    'table_name': 'MANUAL_POST_MATCH_SUMMARY',
    'table_comment': '人工维护-赛后复盘总结表(赛后单次)',
    'source_system': '人工维护/复盘团队',
    'granularity': '赛后单次',
    'integration_method': 'JSON配置',
    'fields': [
        {'name': 'match_id', 'type': 'String', 'comment': '赛事ID(主键)', 'nms_field': ''},
        {'name': 'match_date', 'type': 'String', 'comment': '比赛日期', 'nms_field': ''},
        {'name': 'match_title', 'type': 'String', 'comment': '赛事名称', 'nms_field': ''},
        {'name': 'peak_users', 'type': 'Integer', 'comment': '峰值并发人数', 'nms_field': ''},
        {'name': 'total_traffic_tb', 'type': 'Float', 'comment': '累计总流量(TB)', 'nms_field': ''},
        {'name': 'package_orders', 'type': 'Integer', 'comment': '5G-A专属包订购量', 'nms_field': ''},
        {'name': 'ai_opt_count', 'type': 'Integer', 'comment': '自动优化次数', 'nms_field': ''},
        {'name': 'ai_intercept_count', 'type': 'Integer', 'comment': '隐患主动拦截数', 'nms_field': ''},
        {'name': 'vip_user_count', 'type': 'Integer', 'comment': '重保VIP总人数', 'nms_field': ''},
        {'name': 'vip_satisfaction_rate', 'type': 'Float', 'comment': 'VIP满意度', 'nms_field': ''},
    ]
}

# 汇总所有新表
NEW_TABLES = [
    MAE_15MIN_PERF,
    DSP_15MIN_KQI,
    SEQ_HOURLY_USER,
    SEQ_DAILY_USER_PROFILE,
    SEQ_HOURLY_DEVICE,
    AUTIN_REALTIME_ALARM,
    MANUAL_STATIC_CONFIG,
    MANUAL_STATIC_CAPACITY,
    MANUAL_POST_MATCH_SUMMARY,
]

# ==================== 生成Excel ====================

def create_table_sheet(wb, table_def):
    """为每个表创建Sheet"""
    ws = wb.create_sheet(table_def['table_name'])
    
    # 表头信息
    ws.append(['表名', table_def['table_name']])
    ws.append(['表中文名', table_def['table_comment']])
    ws.append(['数据源系统', table_def['source_system']])
    ws.append(['时间颗粒度', table_def['granularity']])
    ws.append(['数据对接方式', table_def['integration_method']])
    ws.append([])  # 空行
    
    # 字段列表
    headers = ['序号', '字段名', '数据类型', '字段说明', '网管映射字段']
    ws.append(headers)
    
    # 设置表头样式
    for cell in ws[7]:
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
    
    # 添加字段
    for idx, field in enumerate(table_def['fields'], 1):
        ws.append([
            idx,
            field['name'],
            field['type'],
            field['comment'],
            field['nms_field']
        ])
    
    # 调整列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 35
    
    return ws

def create_index_sheet(wb):
    """创建总目录Sheet"""
    ws = wb.create_sheet('00_总目录', 0)
    
    headers = ['序号', '物理表名', '表中文名', '数据源系统', '时间颗粒度', '数据对接方式', '字段数']
    ws.append(headers)
    
    # 设置表头样式
    for cell in ws[1]:
        cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
    
    # 添加表信息
    for idx, table in enumerate(NEW_TABLES, 1):
        ws.append([
            idx,
            table['table_name'],
            table['table_comment'],
            table['source_system'],
            table['granularity'],
            table['integration_method'],
            len(table['fields'])
        ])
    
    # 空行后添加说明
    ws.append([])
    ws.append(['说明'])
    ws.append(['本版本按数据源系统 + 时间颗粒度重新切分底层物理表'])
    ws.append(['共9张物理表，涵盖5大数据源系统'])
    
    # 调整列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 10
    
    return ws

def create_mapping_sheet(wb):
    """创建旧表与新表映射关系Sheet"""
    ws = wb.create_sheet('旧新表映射关系')
    
    headers = ['旧表名(V4)', '旧表用途', '新物理表名', '数据源系统', '切分说明']
    ws.append(headers)
    
    # 设置表头样式
    for cell in ws[1]:
        cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        cell.font = Font(bold=True)
    
    mappings = [
        ['01_Master_Cell_Dim', '网元公参维表', 'MANUAL_CELL_CONFIG', '人工维护/网优团队', '静态配置表'],
        ['02_Master_User_Dim', '用户画像维表', 'SEQ_USER_PROFILE_DAILY', 'SEQ系统(共享层)', '天级画像'],
        ['03_Master_Cell_Perf_Realtime', '基站性能流水', 'MAE_PERF_15MIN', '无线网络管平台 (MAE)', '15分钟性能指标'],
        ['04_P0_User_Active_Hourly', '用户轨迹流', 'SEQ_USER_HOURLY', 'SEQ系统(共享层)', '小时级用户数据'],
        ['05_P1_Venue_KQI_Fact', '场馆KQI指标', 'DSP_KQI_15MIN', '无线智能板 (DSP)', '15分钟KQI'],
        ['06_P1_OAM_Event_Stream', '告警事件流', 'AUTIN_ALARM_REALTIME', '故障告警中心 (AUTIN)', '实时告警'],
        ['07_P2_Zone_User_Agg_Fact', '区域用户聚合', '多表拆分', '多系统', '拆分到SEQ/DSP/人工'],
        ['08_P2_App_Exp_Realtime', 'App体验流水', 'DSP_KQI_15MIN', '无线智能板 (DSP)', '15分钟KQI'],
        ['09_P3_Match_Summary_Fact', '赛后战报', 'MANUAL_POST_MATCH_SUMMARY', '人工维护/复盘团队', '赛后单次'],
        ['10_P3_AI_Opt_Log_Fact', 'AI优化记录', 'MANUAL_POST_MATCH_SUMMARY', '人工维护/复盘团队', '赛后单次'],
        ['11_P3_VIP_Care_Fact', 'VIP核查单', 'MANUAL_POST_MATCH_SUMMARY', '人工维护/复盘团队', '赛后单次'],
    ]
    
    for row in mappings:
        ws.append(row)
    
    # 调整列宽
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    
    return ws

def main():
    print('=' * 70)
    print('苏超智能化指挥中心 - 数据血缘表 V5')
    print('按数据源系统 + 时间颗粒度重新切分底层表格')
    print('=' * 70)
    
    # 创建工作簿
    wb = Workbook()
    wb.remove(wb.active)  # 删除默认Sheet
    
    # 创建总目录
    create_index_sheet(wb)
    print('\n1. 创建总目录')
    
    # 创建各表
    for table_def in NEW_TABLES:
        create_table_sheet(wb, table_def)
        print(f'2. 创建表: {table_def["table_name"]} ({len(table_def["fields"])}个字段)')
    
    # 创建映射关系表
    create_mapping_sheet(wb)
    print('3. 创建旧新表映射关系')
    
    # 保存
    wb.save(OUTPUT_FILE)
    print(f'\n输出文件: {OUTPUT_FILE}')
    print('=' * 70)

if __name__ == '__main__':
    main()
